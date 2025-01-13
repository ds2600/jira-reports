import re
import requests
import pandas as pd
import json
import os
import configparser
import smtplib
import argparse
import logging
from logging.handlers import SysLogHandler
from email.message import EmailMessage
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill

# Parse arguments
parser = argparse.ArgumentParser(description="Generate a Jira report and send it as an attachment")
parser.add_argument("--email", help="Email address to send the report to")
parser.add_argument("--debug", action="store_true", help="Output log messages to console")
parser.add_argument("--plain", action="store_true", help="Outputs non-formatted Excel file")
args = parser.parse_args()

# Retrieve and build configurations
config = configparser.ConfigParser()
config.read("config.ini")

JIRA_BASE_URL = config["credentials"]["JIRA_BASE_URL"]
API_EMAIL = config["credentials"]["API_EMAIL"]
API_KEY = config["credentials"]["API_KEY"]
PROJECT_KEY = config["credentials"]["PROJECT_KEY"]

SMTP_SERVER = config["smtp"]["SMTP_SERVER"]
SMTP_PORT = int(config["smtp"]["SMTP_PORT"])
SMTP_DEBUG = int(config["smtp"].get("SMTP_DEBUG", 0))
FROM_EMAIL = config["smtp"]["FROM_EMAIL"]
REPLY_TO = config["smtp"]["REPLY_TO"]
SUBJECT = config["smtp"]["SUBJECT"]

def setup_logging(logfile, debug=False):
    """
    Configure logging system.
    
    :param logfile: Path to the logfile
    :param debug: Boolean indicating whether to enable debug-level logging
    :return: Configured logger instance
    """
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG if debug else logging.INFO)
    
    log_format = "%(asctime)s %(name)s[%(process)d]: %(levelname)s %(message)s"
    
    file_handler = logging.FileHandler(logfile)
    file_handler.setLevel(logging.DEBUG if debug else logging.INFO)
    file_handler.setFormatter(logging.Formatter(log_format))
    logger.addHandler(file_handler)
    
    if debug:
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.DEBUG)
        console.handler.setFormatter(logging.Formatter(log_format))
        logger.addHandler(console_handler)
        logger.info("Debugging mode enabled")
        
    return logger

# Setup logging
logfile = os.path.join(os.path.dirname(os.path.abspath(__file__)), "script.log")
logger = setup_logging(logfile, debug=args.debug)


# JQL to fetch Epics and related issues
JQL_QUERY = f'issuetype = Epic AND project = {PROJECT_KEY} ORDER BY key ASC' 

def fetch_issues(jql_query):
    """
    Fetch issues from Jira based on the provided JQL query.
    
    :param jql_query: JQL query string to filter issues
    :return: Parsed JSON response contining Jira issues
    :raises: Exception if the request fails
    """
    
    url = f"{JIRA_BASE_URL}/rest/api/3/search"
    headers = {
        "Accept": "application/json",
    }
    auth = (API_EMAIL, API_KEY)
    params = {
        "jql": jql_query,
        "maxResults": 100,  
        "fields": "key,summary,comment, status"
    }
    logger.info(f"Initiating request to fetch issues with JQL: {jql_query}")
    logger.debug(f"Request URL: {url}")
    logger.debug(f"Request headers: {headers}")
    logger.debug(f"Request params: {params}")
    
    
    try:
        response = requests.get(url, headers=headers, auth=auth, params=params)
        logger.debug(f"Response status code: {response.status_code}")
        response.raise_for_status()
        logger.info("Successfully fetched issues from Jira")
        logger.debug(f"Response JSON: {response.json()}")
        return response.json()
    except requests.exceptions.HTTPError as http_err:
        if response.status_code == 401:
            logger.error("Authentication failed: Invalid or expired API key")
        elif response.status_code == 403:
            logger.error("Access denied: You do not have permission to access this resource")
        else:
            logger.error(f"HTTP error occurred: {http_err} - Status Code: {response.status_code}")
        raise
    except requests.exceptions.RequestException as req_err:
        logger.error(f"Request error occurred: {req_err}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error occurred while fetching issues: {e}")
        raise
    

def fetch_most_recent_comment(issue_key):
    """
    Fetches the most recent comment for a given Jira issue.
    
    :param issue_key: The key of hte Jira issue (i.e. "PROJECT-123")
    :return: A tuple containing the comment body and formatted comment date, or None if no comments exist.
    :raises: Exception if the request fails
    """
    
    url = f"{JIRA_BASE_URL}/rest/api/3/issue/{issue_key}/comment"
    headers = {
        "Accept": "application/json",
    }
    auth = (API_EMAIL, API_KEY)
    
    logger.info(f"Fetching most recent comment for issue: {issue_key}")
    logger.debug(f"Request URL: {url}")
    logger.debug(f"Request headers: {headers}")
    
    try:
        response = requests.get(url, headers=headers, auth=auth)
        logger.debug(f"Response status code: {response.status_code}")
        response.raise_for_status()
        
        comments = response.json()["comments"]
        if comments:
            logger.info(f"Found {len(comments)} comments for issue {issue_key}")
            recent_comment = comments[-1]
            comment_body = parse_comment(recent_comment["body"])
            comment_date = recent_comment["created"]
        
            try:
                dt = datetime.strptime(comment_date, '%Y-%m-%dT%H:%M:%S.%f%z')
            except ValueError:
                dt = datetime.strptime(comment_date, '%Y-%m-%dT%H:%M:%S%z')
            
            formatted_comment_date = dt.strftime('%Y-%m-%d %H:%M:%S %z')
            logger.debug(f"Most recent comment date: {formatted_comment_date}")
            return comment_body, formatted_comment_date
            
        logger.info(f"No comments found for issue {issue_key}")
        return " ", None
        
    except requests.exceptions.HTTPError as http_err:
        if response.status_code == 401:
            logger.error("Authentication failed: Invalid or expired API key.")
        elif response.status_code == 403:
            logger.error("Access denied: You do not have permission to access this resource.")
        elif response.status_code == 404:
            logger.error(f"Issue {issue_key} not found.")
        else:
            logger.error(f"HTTP error occurred: {http_err} - Status Code: {response.status_code}")
        raise
    except requests.exceptions.RequestException as req_err:
        logger.error(f"Request error occurred: {req_err}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error while fetching comments for issue {issue_key}: {e}")
        raise

def parse_comment(comment_json):
    """
    Parses a Jira comment into human readable format
    
    :param comment_json: JSON object containing the comment
    :return: A string containing the parsed comment
    """
    
    def parse_block(block):
        """
        Recursively parse a block of JSON and return its text representation.
        
        :param block: A JSON object representing a block of content
        :return: A string representation of the blcoks content
        """
        if block["type"] == "paragraph":
            logger.debug("Parsing paragraph block")
            paragraph_text = []
            for content in block.get("content", []):
                if content["type"] == "text":
                    paragraph_text.append(content["text"])
            return " ".join(paragraph_text)
        elif block["type"] == "listItem":
            logger.debug("Parsing list item block")
            bullet_text = []
            for content in block.get("content", []):
                if content["type"] == "paragraph":
                    bullet_text.append(parse_block(content))
            return f"- {' '.join(bullet_text)}"
        elif block["type"] == "orderedList" or block["type"] == "bulletList":
            logger.debug(f"Parsing {block['type']} block")
            list_text = []
            for item in block.get("content", []):
                if item["type"] == "listItem":
                    list_text.append(parse_block(item))
            return "\n".join(list_text)
        
        logger.warning(f"Unhandled block type: {block.get('type')}")
        return ""  # Fallback for unhandled block types

    if isinstance(comment_json, dict) and comment_json.get("type") == "doc":
        logger.info("Parsing a valid comment document")
        text = []
        for block in comment_json.get("content", []):
            parsed_block = parse_block(block)
            if parsed_block:
                text.append(parsed_block)
        parsed_text = "\n".join(text)
        logger.debug(f"Parsed comment text: {parsed_text}")
        return parsed_text
        
    logger.warning("Invalid or unsupported comment JSON structure")
    return json.dumps(comment_json, indent=2)  # Return formatted JSON for unhandled structures

def generate_excel_report(data, script_dir):
    """
    Generates the Excel report from the collected data.

    :param data: The list of issues and tasks.
    :param script_dir: The directory to save the file.
    :return: The path to the generated Excel file.
    """
    logger.info("Preparing data for Excel report")
    excel_data = []
    for epic_key, group in pd.DataFrame(data).groupby("Epic Key"):
        epic_summary = group.iloc[0]["Epic Summary"]
        excel_data.append([f"{epic_summary} ({epic_key})", "", "", "", "", ""])

        # Sort child issues and sub-tasks
        status_order = {"In Progress": 1, "Waiting": 2, "For approval": 3, "New": 4, "Done": 5}
        child_issues = group[group["Type"] == "Task"].copy()
        sub_tasks = group[group["Type"] == "Sub-Task"].copy()

        logger.debug("Sorting rows for epic")
        child_issues.sort_values(
            by=["Child Status", "Comment Date"],
            key=lambda col: col.map(status_order).fillna(5) if col.name == "Child Status" else col,
            ascending=[True, False],
            inplace=True,
        )
        sub_tasks.sort_values(
            by=["Child Status", "Comment Date", "Parent"],
            key=lambda col: col.map(status_order).fillna(5) if col.name == "Child Status" else col,
            ascending=[True, False, True],
            inplace=True,
        )

        # Combine sorted rows
        for _, child_row in child_issues.iterrows():
            excel_data.append([
                "",
                child_row["Child Key"],
                child_row["Child Summary"],
                child_row["Child Status"],
                child_row["Comment Date"],
                child_row["Most Recent Comment"],
                child_row.get("Indent", 0),
            ])
            child_key = child_row["Child Key"]
            sub_task_rows = sub_tasks[sub_tasks["Parent"] == child_key]
            for _, sub_row in sub_task_rows.iterrows():
                excel_data.append([
                    "",
                    sub_row["Child Key"],
                    sub_row["Child Summary"],
                    sub_row["Child Status"],
                    sub_row["Comment Date"],
                    sub_row["Most Recent Comment"],
                    sub_row.get("Indent", 1),
                ])

        excel_data.append(["§", "", "", "", "", "", ""])  # Spacer row

    # Save the Excel file
    timestamp = datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
    filename = os.path.join(script_dir, f"jira_report_{timestamp}.xlsx")
    logger.info(f"Saving Excel report to {filename}")
    pd.DataFrame(excel_data).to_excel(filename, index=False, header=False)
    return filename

def format_excel_file(filename):
    """
    Formats an Excel file for readable and repeatable Jira issue report
    
    :param filename: The apth to the Excel file
    """
    
    try:
        logger.info(f"Loading workbook: {filename}")
        wb = load_workbook(filename)
        ws = wb.active
        
        header_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')  # Light steel blue
        header_font = Font(bold=True, size=11, color='000000')
        header_alignment = Alignment(horizontal='left', vertical='center')
        header_border = Border(
            bottom=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )
        
        done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Pastel green
        sub_task_spacer_fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")  # Light Gray    
        epic_header_pattern = re.compile(r'.+\(NOOPT-\d+\)$')
    
        logger.debug("Setting column dimensions")
        column_widths = {'A': 2, 'B': 15, 'C': 40, 'D': 20, 'E': 25, 'F': 100}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        def apply_header_style(cell):
            """Applies header styles to a cell."""
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        def apply_task_style(row):
            """Applies styles for a task row."""
            ws.merge_cells(f'A{row}:B{row}')
            task_cell = ws[f'A{row}']
            task_cell.alignment = header_alignment

        def apply_done_style(row, indent):
            """Applies styles for a row marked as 'Done'."""
            if indent == 0:
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'].fill = done_fill
            else:
                for col in ['B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'].fill = done_fill
                
        logger.debug("Formatting rows")
        current_row = 1

        while current_row <= ws.max_row:
            first_cell = ws[f'A{current_row}']
            cell_value = str(first_cell.value) if first_cell.value else ""
            row_key = str(ws[f'B{current_row}'].value)
            logger.debug(f"Processing row {current_row}: {cell_value}")

            if epic_header_pattern.match(cell_value): 
                logger.debug(f'Found header row: {current_row}')
                ws.merge_cells(f'A{current_row}:F{current_row}')
                apply_header_style(ws[f'A{current_row}'])
            else:
                if ws[f'G{current_row}'].value is not None:
                    indent_level = int(ws[f'G{current_row}'].value)
                else:
                    indent_level = 1
                    row_key = ""

                if ws[f'A{current_row}'].value == "§":
                    indent_level = 3
                    
                logger.debug(f'Indent level: {indent_level} Row: {current_row}')
                
                if indent_level == 0:  # Task
                    logger.debug(f'Task row: {current_row}')
                    task_cell = ws[f'A{current_row}']
                    apply_task_style(current_row)
                    task_cell.value = row_key
                   
                elif indent_level == 1:  # Sub-Task
                    logger.debug(f'Sub-Task row: {current_row}')
                    spacer_cell = ws[f'A{current_row}']
                    spacer_cell.value = ""
                    spacer_cell.fill = sub_task_spacer_fill
                    sub_task_cell = ws[f'B{current_row}']
                    sub_task_cell.value = row_key
                    sub_task_cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    logger.debug(f"Skipping spacer row: {current_row}")
                    spacer_cell = ws[f'A{current_row}']
                    spacer_cell.value = ""
                    spacer_cell.fill = PatternFill(fill_type=None)
                
                for col in ['C', 'D', 'E', 'F']:
                    cell = ws[f'{col}{current_row}']
                    cell.alignment = Alignment(
                        wrap_text=True, 
                        vertical='top',
                        horizontal='left',
                        indent=indent_level
                    )
                    
                    if ws[f'D{current_row}'].value and ws[f'D{current_row}'].value.strip() == "Done":
                        logger.debug(f"Applying 'Done' style to row: {current_row}")
                        apply_done_style(current_row, indent_level)
            current_row += 1

        logger.info("Deleting unused columns")
        ws.delete_cols(7)
    
        logger.info(f"Saving formatted workbook: {filename}")
        wb.save(filename)
    except Exception as e:
        logger.error(f"Error occurred during file formatting: {e}")
        raise

def process_child_issue(data, epic_key, epic_summary, child_issue):
    """
    Processes a child issue, including its subtasks and appends it's results to data
    """
    child_key = child_issue["key"]
    child_summary = child_issue["fields"]["summary"]
    child_status = child_issue["fields"]["status"]["name"]
    
    logger.info(f"Processing child issue: {child_key}")
    most_recent_comment, comment_date = fetch_most_recent_comment(child_key)
    
    data.append({
        "Epic Key": epic_key,
        "Epic Summary": epic_summary,
        "Child Key": child_key,
        "Child Summary": child_summary,
        "Child Status": child_status,
        "Comment Date": comment_date,
        "Most Recent Comment": most_recent_comment,
        "Type": "Task",
        "Indent": 0,
    })
    
    sub_tasks = fetch_issues(f'"parent" = {child_key}')
    if sub_tasks["issues"]:
        for sub_task in sub_tasks["issues"]:
            process_sub_task(data, epic_key, epic_summary, child_key, sub_task)

    
def process_sub_task(data, epic_key, epic_summary, parent_key, sub_task):
    """
    Processes a sub-task and appends the results to the data list.
    """
    sub_key = sub_task["key"]
    sub_summary = sub_task["fields"]["summary"]
    sub_status = sub_task["fields"]["status"]["name"]

    logger.debug(f"Processing sub-task: {sub_key} - {sub_summary}")
    sub_comment, sub_comment_date = fetch_most_recent_comment(sub_key)

    data.append({
        "Epic Key": epic_key,
        "Epic Summary": epic_summary,
        "Child Key": sub_key,
        "Child Summary": sub_summary,
        "Child Status": sub_status,
        "Comment Date": sub_comment_date,
        "Most Recent Comment": sub_comment,
        "Type": "Sub-Task",
        "Parent": parent_key,
        "Indent": 1,
    })

def send_email(recipient, filename):
    """
    Send an email with the specified attachemnt to the recipient
    
    :param recipient: Recipient email address
    :param filename: Path to file to attached
    :raises: Excpetion if email fails to send
    """
    
    logger.info(f"Preparing to send email to {recipient} with attachment {filename}")
    try:
        logger.debug("Setting up the email message")
        msg = EmailMessage()
        msg["From"] = FROM_EMAIL
        msg["To"] = recipient
        msg["Reply-To"] = REPLY_TO
        msg["Subject"] = SUBJECT
        msg.set_content("Please see the attached Jira report.")

        logger.debug(f"Attaching file: {filename}")
        with open(filename, "rb") as f:
            file_data = f.read()
            file_name = os.path.basename(filename)
            msg.add_attachment(file_data, maintype="application", subtype="octet-stream", filename=file_name)
    
        logger.info(f"Connecting to SMTP server: {SMTP_SERVER}:{SMTP_PORT}")
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            logger.info(f"Sending email to {recipient}")
            server.send_message(msg)
        logger.info(f"Email sent successfully to {recipient}")
    except FileNotFoundError:
        logger.error(f"Attachment file not found: {filename}")
        raise
    except smtplib.SMTPException as smtp_err:
        logger.error(f"SMTP error occurred while sending email to {recipient}: {smtp_err}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error occurred while sending email to {recipient}: {e}")
        raise

def main():
    logger.info("Script started")
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        logger.info(f"Working directory: {script_dir}")
    
        logger.info("Fetching Epics...")
        print("[ ] Fetching epics...[ ]", end="\r", flush=True)
        epics = fetch_issues(JQL_QUERY)
        
        if not epics["issues"]:
            logger.warning("No epics found, exiting")
            print("[X] Fetching epics...[0/0]")
            return
        
        total_epics = len(epics["issues"])
        print(f"[ ] Fetching epics...[0/{total_epics}]", end="\r", flush=True)
        
        data = []
        for i, epic in enumerate(epics["issues"], start=1):
            epic_key = epic["key"]
            epic_summary = epic["fields"]["summary"]
            logger.info(f"Processing Epic: {epic_key} - {epic_summary}")
            
            print(f"[ ] Fetching epics...[{i}/{total_epics}]", end="\r", flush=True)
            
            child_issues = fetch_issues(f'"Epic Link" = {epic_key}')
            if child_issues["issues"]:
                for child in child_issues["issues"]:
                    process_child_issue(data, epic_key, epic_summary, child)
                   
            else:
                logger.info(f"No child issue found for epic: {epic_key}")
                data.append({
                    "Epic Key": epic_key,
                    "Epic Summary": epic_summary,
                    "Child Key": "",
                    "Child Summary": "No tasks created",
                    "Comment Date": "",
                    "Most Recent Comment": "",
                })
        print(f"[X] Fetching epics...[{total_epics}/{total_epics}]")
            
        logger.info("Generating Excel report")
        print("[ ] Generating report...", end="\r", flush=True)
        filename = generate_excel_report(data, script_dir)
        print("[X] Generating report...")
        logger.info("Report generation complete")
        
        if not getattr(args, 'plain', False): 
            print("[ ] Formatting report...", end="\r", flush=True)
            logger.info("Applying formatting...")
            format_excel_file(filename)
            print("[X] Formatting report...")
            logger.info("Formatting complete")

        print(f"[X] Report saved as '{filename}'")
        logger.info(f"Report saved as '{filename}'")
        return filename
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}")
        raise

if __name__ == "__main__":
    report_filename = main()
    if args.email:
        recipient = args.email
        print(f"Sending report to {recipient}...")
        send_email(recipient, report_filename)

